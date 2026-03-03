"""
Bulk upload for items and companies.

Two-pass flow for items:
  POST /bulk/items/preview   — parse & validate, return new/duplicate/error rows
  POST /bulk/items/confirm   — commit approved rows with per-duplicate action

Single-pass for companies (no meaningful duplicates beyond name):
  POST /bulk/companies/preview
  POST /bulk/companies/confirm

Template downloads:
  GET  /bulk/items/template
  GET  /bulk/companies/template
"""

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
import os
from typing import Literal, Optional
from io import BytesIO
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import SessionLocal
from app.auth import get_current_user, require_super_admin
from app.models.item import Item
from app.models.company import Company
from app.models.category import Category
from app.models.tag import Tag
from app.models.user import User

router = APIRouter(prefix="/bulk", tags=["Bulk Upload"])

MAX_ROWS = 5000


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean(val) -> str:
    """Coerce cell value to stripped string."""
    if val is None:
        return ""
    return str(val).strip()


def _bool_cell(val) -> bool:
    return _clean(val).lower() in ("yes", "true", "1", "y")


def _get_or_create_company(db: Session, name: str, is_manufacturer=False, is_supplier=False) -> Optional[Company]:
    if not name:
        return None
    name = name.strip()
    company = db.query(Company).filter(Company.name.ilike(name)).first()
    if not company:
        company = Company(
            name=name,
            is_manufacturer=is_manufacturer,
            is_supplier=is_supplier,
        )
        db.add(company)
        db.flush()
    else:
        # Update flags if newly needed
        changed = False
        if is_manufacturer and not company.is_manufacturer:
            company.is_manufacturer = True
            changed = True
        if is_supplier and not company.is_supplier:
            company.is_supplier = True
            changed = True
        if changed:
            db.flush()
    return company


def _style_header_row(ws, row: int, col_count: int):
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_example_row(ws, row: int, col_count: int):
    fill = PatternFill("solid", fgColor="EBF4FF")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(italic=True, color="555555", size=9)
        cell.alignment = Alignment(wrap_text=True)


def _auto_width(ws, min_width=12, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ══════════════════════════════════════════════════════════════════════════════
# ITEMS
# ══════════════════════════════════════════════════════════════════════════════

ITEM_COLUMNS = [
    ("name",          "Name *",           "Fire Extinguisher CO2 5kg"),
    ("catalogue_nr",  "Catalogue Nr",     "FE-CO2-5"),
    ("unit",          "Unit *",           "pcs"),
    ("category",      "Category *",       "Safety"),
    ("desc_short",    "Short Description","CO2 extinguisher, 5kg"),
    ("desc_long",     "Full Description", "For use in engine room and electrical fires"),
    ("manufacturer",  "Manufacturer",     "Amerex"),
    ("supplier",      "Supplier",         "MarineStore Ltd"),
    ("tags",          "Tags",             "Safety, Critical"),
    ("image_path",    "Image Path",       "media/items/fe-co2-5.jpg"),
]

COL_IDX = {col[0]: i + 1 for i, col in enumerate(ITEM_COLUMNS)}


@router.get("/items/template")
def download_items_template(_: User = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    ws.row_dimensions[1].height = 36

    # Notes row
    notes = [
        "* Required",
        "Must match an existing catalogue number exactly if updating",
        "* Required  e.g. pcs / L / kg / m",
        "* Must match existing category name exactly",
        "Optional short summary",
        "Optional detailed description",
        "Created automatically if not found",
        "Created automatically if not found",
        "Comma-separated existing tag names",
        "Path to existing file on server e.g. media/items/photo.jpg",
    ]

    for i, (_, header, example) in enumerate(ITEM_COLUMNS, start=1):
        ws.cell(1, i, header)
        ws.cell(2, i, example)
        ws.cell(3, i, notes[i - 1])

    _style_header_row(ws, 1, len(ITEM_COLUMNS))
    _style_example_row(ws, 2, len(ITEM_COLUMNS))

    note_fill = PatternFill("solid", fgColor="FFF9E6")
    for col in range(1, len(ITEM_COLUMNS) + 1):
        cell = ws.cell(3, col)
        cell.fill = note_fill
        cell.font = Font(italic=True, color="888800", size=8)
        cell.alignment = Alignment(wrap_text=True)

    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"
    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items_template.xlsx"},
    )


class ItemRowPreview(BaseModel):
    row: int
    status: Literal["new", "duplicate", "error"]
    name: str
    catalogue_nr: str
    unit: str
    category: str
    desc_short: str
    desc_long: str
    manufacturer: str
    supplier: str
    tags: str
    image_path: str
    error: Optional[str] = None
    existing_id: Optional[int] = None


class ItemPreviewResponse(BaseModel):
    total: int
    new_count: int
    duplicate_count: int
    error_count: int
    rows: list[ItemRowPreview]


@router.post("/items/preview", response_model=ItemPreviewResponse)
def preview_items_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_super_admin),
):
    contents = file.file.read()
    try:
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid Excel file")

    ws = wb.active
    rows_iter = iter(ws.iter_rows(values_only=True))

    # Skip header rows (row 1 = headers, row 2 = example, row 3 = notes)
    for _ in range(3):
        try:
            next(rows_iter)
        except StopIteration:
            break

    # Cache for validation
    categories = {c.name.lower(): c for c in db.query(Category).all()}
    tags_map = {t.name.lower(): t for t in db.query(Tag).all()}

    results: list[ItemRowPreview] = []
    seen_in_file: set[str] = set()  # track duplicates within the file itself

    for row_num, row in enumerate(rows_iter, start=4):
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # skip blank rows
        if len(results) >= MAX_ROWS:
            break

        name = _clean(row[COL_IDX["name"] - 1])
        catalogue_nr = _clean(row[COL_IDX["catalogue_nr"] - 1])
        unit = _clean(row[COL_IDX["unit"] - 1])
        category_name = _clean(row[COL_IDX["category"] - 1])
        desc_short = _clean(row[COL_IDX["desc_short"] - 1])
        desc_long = _clean(row[COL_IDX["desc_long"] - 1])
        manufacturer_name = _clean(row[COL_IDX["manufacturer"] - 1])
        supplier_name = _clean(row[COL_IDX["supplier"] - 1])
        tags_raw = _clean(row[COL_IDX["tags"] - 1])
        image_path = _clean(row[COL_IDX["image_path"] - 1])

        errors = []
        if not name:
            errors.append("Name is required")
        if not unit:
            errors.append("Unit is required")
        if not category_name:
            errors.append("Category is required")
        elif category_name.lower() not in categories:
            errors.append(f"Category '{category_name}' not found")

        if errors:
            results.append(ItemRowPreview(
                row=row_num, status="error",
                name=name, catalogue_nr=catalogue_nr, unit=unit,
                category=category_name, desc_short=desc_short, desc_long=desc_long,
                manufacturer=manufacturer_name, supplier=supplier_name,
                tags=tags_raw, image_path=image_path, error="; ".join(errors),
            ))
            continue

        # Duplicate detection: match on (name, catalogue_nr) or just name if no catalogue_nr
        existing = None
        if catalogue_nr:
            existing = db.query(Item).filter(
                Item.name.ilike(name),
                Item.catalogue_nr.ilike(catalogue_nr),
            ).first()
        if not existing:
            existing = db.query(Item).filter(Item.name.ilike(name)).first()

        # Also check for duplicates within this file
        file_key = f"{name.lower()}|{catalogue_nr.lower()}"
        if file_key in seen_in_file:
            results.append(ItemRowPreview(
                row=row_num, status="error",
                name=name, catalogue_nr=catalogue_nr, unit=unit,
                category=category_name, desc_short=desc_short, desc_long=desc_long,
                manufacturer=manufacturer_name, supplier=supplier_name,
                tags=tags_raw, image_path=image_path, error="Duplicate row within this file",
            ))
            continue
        seen_in_file.add(file_key)

        status = "duplicate" if existing else "new"
        results.append(ItemRowPreview(
            row=row_num, status=status,
            name=name, catalogue_nr=catalogue_nr, unit=unit,
            category=category_name, desc_short=desc_short, desc_long=desc_long,
            manufacturer=manufacturer_name, supplier=supplier_name,
            tags=tags_raw, image_path=image_path,
            existing_id=existing.id if existing else None,
        ))

    wb.close()

    new_count = sum(1 for r in results if r.status == "new")
    dup_count = sum(1 for r in results if r.status == "duplicate")
    err_count = sum(1 for r in results if r.status == "error")

    return ItemPreviewResponse(
        total=len(results),
        new_count=new_count,
        duplicate_count=dup_count,
        error_count=err_count,
        rows=results,
    )


class ItemConfirmRow(BaseModel):
    row: int
    action: Literal["create", "update", "skip"]
    name: str
    catalogue_nr: str
    unit: str
    category: str
    desc_short: str
    desc_long: str
    manufacturer: str
    supplier: str
    tags: str
    image_path: str
    existing_id: Optional[int] = None


class ItemConfirmRequest(BaseModel):
    rows: list[ItemConfirmRow]


class ItemConfirmResponse(BaseModel):
    created: int
    updated: int
    skipped: int
    errors: list[str]


@router.post("/items/confirm", response_model=ItemConfirmResponse)
def confirm_items_upload(
    data: ItemConfirmRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_super_admin),
):
    categories = {c.name.lower(): c for c in db.query(Category).all()}
    tags_map = {t.name.lower(): t for t in db.query(Tag).all()}

    created = updated = skipped = 0
    errors = []

    for row in data.rows:
        if row.action == "skip":
            skipped += 1
            continue

        category = categories.get(row.category.lower())
        if not category:
            errors.append(f"Row {row.row}: category '{row.category}' not found")
            continue

        manufacturer = _get_or_create_company(db, row.manufacturer, is_manufacturer=True) if row.manufacturer else None
        supplier = _get_or_create_company(db, row.supplier, is_supplier=True) if row.supplier else None

        tag_names = [t.strip().lower() for t in row.tags.split(",") if t.strip()]
        item_tags = [tags_map[n] for n in tag_names if n in tags_map]

        if row.action == "update" and row.existing_id:
            item = db.get(Item, row.existing_id)
            if not item:
                errors.append(f"Row {row.row}: item id {row.existing_id} not found")
                continue
            item.name = row.name
            item.catalogue_nr = row.catalogue_nr or None
            item.unit = row.unit
            item.category_id = category.id
            item.desc_short = row.desc_short or None
            item.desc_long = row.desc_long or None
            item.manufacturer_id = manufacturer.id if manufacturer else None
            item.supplier_id = supplier.id if supplier else None
            item.tags = item_tags
            if row.image_path:
                if os.path.exists(row.image_path):
                    item.image_path = row.image_path
                else:
                    errors.append(f"Row {row.row}: image file not found: {row.image_path}")
            updated += 1

        elif row.action == "create":
            item = Item(
                name=row.name,
                catalogue_nr=row.catalogue_nr or None,
                unit=row.unit,
                category_id=category.id,
                desc_short=row.desc_short or None,
                desc_long=row.desc_long or None,
                manufacturer_id=manufacturer.id if manufacturer else None,
                supplier_id=supplier.id if supplier else None,
                created_by=current_user.id,
                is_active=True,
            )
            if row.image_path:
                if os.path.exists(row.image_path):
                    item.image_path = row.image_path
                else:
                    errors.append(f"Row {row.row}: image file not found: {row.image_path}")
            item.tags = item_tags
            db.add(item)
            created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Database error during commit: {str(e)}")

    return ItemConfirmResponse(created=created, updated=updated, skipped=skipped, errors=errors)


# ══════════════════════════════════════════════════════════════════════════════
# COMPANIES
# ══════════════════════════════════════════════════════════════════════════════

COMPANY_COLUMNS = [
    ("name",            "Name *",           "MarineStore Ltd"),
    ("email",           "Email",            "orders@marinestore.com"),
    ("phone",           "Phone",            "+44 20 1234 5678"),
    ("website",         "Website",          "https://marinestore.com"),
    ("is_supplier",     "Is Supplier",      "Yes"),
    ("is_manufacturer", "Is Manufacturer",  "No"),
    ("comments",        "Comments",         "Main UK supplier"),
]

CO_COL_IDX = {col[0]: i + 1 for i, col in enumerate(COMPANY_COLUMNS)}


@router.get("/companies/template")
def download_companies_template(_: User = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    ws.row_dimensions[1].height = 36

    notes = [
        "* Required — must be unique",
        "Optional",
        "Optional",
        "Optional",
        "Yes / No",
        "Yes / No",
        "Optional notes",
    ]

    for i, (_, header, example) in enumerate(COMPANY_COLUMNS, start=1):
        ws.cell(1, i, header)
        ws.cell(2, i, example)
        ws.cell(3, i, notes[i - 1])

    _style_header_row(ws, 1, len(COMPANY_COLUMNS))
    _style_example_row(ws, 2, len(COMPANY_COLUMNS))

    note_fill = PatternFill("solid", fgColor="FFF9E6")
    for col in range(1, len(COMPANY_COLUMNS) + 1):
        cell = ws.cell(3, col)
        cell.fill = note_fill
        cell.font = Font(italic=True, color="888800", size=8)

    ws.freeze_panes = "A4"
    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=companies_template.xlsx"},
    )


class CompanyRowPreview(BaseModel):
    row: int
    status: Literal["new", "duplicate", "error"]
    name: str
    email: str
    phone: str
    website: str
    is_supplier: bool
    is_manufacturer: bool
    comments: str
    error: Optional[str] = None
    existing_id: Optional[int] = None


class CompanyPreviewResponse(BaseModel):
    total: int
    new_count: int
    duplicate_count: int
    error_count: int
    rows: list[CompanyRowPreview]


@router.post("/companies/preview", response_model=CompanyPreviewResponse)
def preview_companies_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_super_admin),
):
    contents = file.file.read()
    try:
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid Excel file")

    ws = wb.active
    rows_iter = iter(ws.iter_rows(values_only=True))
    for _ in range(3):
        try:
            next(rows_iter)
        except StopIteration:
            break

    results: list[CompanyRowPreview] = []
    seen: set[str] = set()

    for row_num, row in enumerate(rows_iter, start=4):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        if len(results) >= MAX_ROWS:
            break

        name = _clean(row[CO_COL_IDX["name"] - 1])
        email = _clean(row[CO_COL_IDX["email"] - 1])
        phone = _clean(row[CO_COL_IDX["phone"] - 1])
        website = _clean(row[CO_COL_IDX["website"] - 1])
        is_supplier = _bool_cell(row[CO_COL_IDX["is_supplier"] - 1])
        is_manufacturer = _bool_cell(row[CO_COL_IDX["is_manufacturer"] - 1])
        comments = _clean(row[CO_COL_IDX["comments"] - 1])

        if not name:
            results.append(CompanyRowPreview(
                row=row_num, status="error",
                name="", email=email, phone=phone, website=website,
                is_supplier=is_supplier, is_manufacturer=is_manufacturer,
                comments=comments, error="Name is required",
            ))
            continue

        if name.lower() in seen:
            results.append(CompanyRowPreview(
                row=row_num, status="error",
                name=name, email=email, phone=phone, website=website,
                is_supplier=is_supplier, is_manufacturer=is_manufacturer,
                comments=comments, error="Duplicate name within this file",
            ))
            continue
        seen.add(name.lower())

        existing = db.query(Company).filter(Company.name.ilike(name)).first()
        results.append(CompanyRowPreview(
            row=row_num, status="duplicate" if existing else "new",
            name=name, email=email, phone=phone, website=website,
            is_supplier=is_supplier, is_manufacturer=is_manufacturer,
            comments=comments,
            existing_id=existing.id if existing else None,
        ))

    wb.close()
    return CompanyPreviewResponse(
        total=len(results),
        new_count=sum(1 for r in results if r.status == "new"),
        duplicate_count=sum(1 for r in results if r.status == "duplicate"),
        error_count=sum(1 for r in results if r.status == "error"),
        rows=results,
    )


class CompanyConfirmRow(BaseModel):
    row: int
    action: Literal["create", "update", "skip"]
    name: str
    email: str
    phone: str
    website: str
    is_supplier: bool
    is_manufacturer: bool
    comments: str
    existing_id: Optional[int] = None


class CompanyConfirmRequest(BaseModel):
    rows: list[CompanyConfirmRow]


class CompanyConfirmResponse(BaseModel):
    created: int
    updated: int
    skipped: int
    errors: list[str]


@router.post("/companies/confirm", response_model=CompanyConfirmResponse)
def confirm_companies_upload(
    data: CompanyConfirmRequest,
    db: Session = Depends(get_db),
    _: User = Depends(require_super_admin),
):
    created = updated = skipped = 0
    errors = []

    for row in data.rows:
        if row.action == "skip":
            skipped += 1
            continue

        if row.action == "update" and row.existing_id:
            company = db.get(Company, row.existing_id)
            if not company:
                errors.append(f"Row {row.row}: company id {row.existing_id} not found")
                continue
            company.name = row.name
            company.email = row.email or None
            company.phone = row.phone or None
            company.website = row.website or None
            company.is_supplier = row.is_supplier
            company.is_manufacturer = row.is_manufacturer
            company.comments = row.comments or None
            updated += 1

        elif row.action == "create":
            company = Company(
                name=row.name,
                email=row.email or None,
                phone=row.phone or None,
                website=row.website or None,
                is_supplier=row.is_supplier,
                is_manufacturer=row.is_manufacturer,
                comments=row.comments or None,
                is_active=True,
            )
            db.add(company)
            created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Database error: {str(e)}")

    return CompanyConfirmResponse(created=created, updated=updated, skipped=skipped, errors=errors)
