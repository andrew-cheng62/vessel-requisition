from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from datetime import datetime
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from io import BytesIO
from math import ceil

from app.database import SessionLocal
from app.models.requisition import Requisition
from app.models.requisition_item import RequisitionItem
from app.models.user import User
from app.auth import get_current_user, require_captain
from app.schemas.requisition import RequisitionCreate, RequisitionUpdate, RequisitionOut, PaginatedRequisitions

ALLOWED_STATUS_TRANSITIONS = {
    "draft": {"rfq_sent", "cancelled"},
    "rfq_sent": {"ordered", "cancelled"},
    "ordered": {"partially_received", "received", "cancelled"},
    "partially_received": {"received"},
    "received": set(),
    "cancelled": set(),
}

def is_closed(status: str) -> bool:
    return status in ["received", "cancelled"]

def can_edit(req, user):
    if is_closed(req.status):
        return False
    if req.status == "draft":
        return True
    return user.role in ("captain", "super_admin")

def can_change_status(req, user, new_status: str) -> bool:
    if user.role not in ("captain", "super_admin"):
        return False
    return new_status in ALLOWED_STATUS_TRANSITIONS.get(req.status, set())

def can_receive(req, user):
    return (
        user.role in {"captain", "crew", "super_admin"}
        and req.status in {"ordered", "partially_received"}
    )

def can_delete(req, user):
    if user.role not in ("captain", "super_admin"):
        return False
    return req.status in {"draft", "cancelled"}

router = APIRouter(prefix="/requisitions", tags=["Requisitions"])

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_req_or_404(req_id: int, vessel_id: int, db: Session) -> Requisition:
    req = db.query(Requisition).filter(
        Requisition.id == req_id,
        Requisition.vessel_id == vessel_id
    ).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    return req


@router.post("/", response_model=RequisitionOut)
def create_requisition(
    data: RequisitionCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not data.items:
        raise HTTPException(400, "Requisition must contain at least one item")

    try:
        requisition = Requisition(
            supplier_id=data.supplier_id,
            status="draft",
            created_by=user.id,
            created_at=datetime.utcnow(),
            notes=data.notes,
            vessel_id=user.vessel_id,   # scope to vessel
        )
        db.add(requisition)
        db.flush()

        for row in data.items:
            db.add(RequisitionItem(
                requisition_id=requisition.id,
                item_id=row.item_id,
                quantity=row.quantity,
                received_qty=0,
            ))

        db.commit()
        db.refresh(requisition)
        return requisition
    except Exception as e:
        db.rollback()
        raise


@router.get("/", response_model=PaginatedRequisitions)
def list_requisitions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: str | None = None,
    supplier_id: int | None = None,
    active_only: bool = Query(True),   # ← default True
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):

    CLOSED_STATUSES = ["received", "cancelled"]

    q = (
        db.query(Requisition)
        .options(
            joinedload(Requisition.supplier),
            joinedload(Requisition.items),
        )
        .filter(Requisition.vessel_id == current_user.vessel_id)
    )

    if status:
        q = q.filter(Requisition.status == status)
    elif active_only:
        # No explicit status filter + active = hide closed
        q = q.filter(Requisition.status.notin_(CLOSED_STATUSES))

    if supplier_id:
        q = q.filter(Requisition.supplier_id == supplier_id)

    total = q.count()
    items = (
        q.order_by(Requisition.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {"items": items, "total": total, "page": page, "page_size": page_size, "pages": ceil(total / page_size)}


@router.get("/{req_id}", response_model=RequisitionOut)
def get_requisition(
    req_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    req = (
        db.query(Requisition)
        .options(
            joinedload(Requisition.supplier),
            joinedload(Requisition.items).joinedload(RequisitionItem.item),
        )
        .filter(Requisition.id == req_id, Requisition.vessel_id == current_user.vessel_id)
        .first()
    )
    if not req:
        raise HTTPException(404, "Requisition not found")
    return req


@router.post("/{req_id}/status", response_model=RequisitionOut)
def change_status(
    req_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    req = get_req_or_404(req_id, current_user.vessel_id, db)

    new_status = data.get("status")
    if not new_status:
        raise HTTPException(400, "Status required")
    if not can_change_status(req, current_user, new_status):
        raise HTTPException(403, "Invalid status transition")

    if new_status == "ordered":
        req.ordered_at = datetime.utcnow()

    req.status = new_status
    db.commit()

    return db.query(Requisition).options(
        joinedload(Requisition.supplier),
        joinedload(Requisition.items).joinedload(RequisitionItem.item),
    ).filter(Requisition.id == req_id).first()


@router.put("/{req_id}", response_model=RequisitionOut)
def edit_requisition(
    req_id: int,
    data: RequisitionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    req = (
        db.query(Requisition)
        .options(joinedload(Requisition.items))
        .filter(Requisition.id == req_id, Requisition.vessel_id == current_user.vessel_id)
        .first()
    )
    if not req:
        raise HTTPException(404, "Requisition not found")
    if not can_edit(req, current_user):
        raise HTTPException(403, "Editing not allowed")

    if data.supplier_id is not None:
        req.supplier_id = data.supplier_id
    if data.notes is not None:
        req.notes = data.notes
    if data.items is not None:
        req.items.clear()
        db.flush()
        for row in data.items:
            req.items.append(RequisitionItem(item_id=row.item_id, quantity=row.quantity, received_qty=0))

    db.commit()

    return db.query(Requisition).options(
        joinedload(Requisition.supplier),
        joinedload(Requisition.items).joinedload(RequisitionItem.item),
    ).filter(Requisition.id == req_id).first()


@router.post("/{req_id}/items", response_model=RequisitionOut)
def add_item_to_requisition(
    req_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    req = (
        db.query(Requisition)
        .options(joinedload(Requisition.items))
        .filter(Requisition.id == req_id, Requisition.vessel_id == current_user.vessel_id)
        .first()
    )
    if not req:
        raise HTTPException(404)
    if req.status != "draft":
        raise HTTPException(403, "Can only add items in draft")

    item_id = data.get("item_id")
    qty = data.get("quantity", 1)
    if not item_id:
        raise HTTPException(400, "item_id required")

    existing = next((i for i in req.items if i.item_id == item_id), None)
    if existing:
        existing.quantity += qty
    else:
        req.items.append(RequisitionItem(item_id=item_id, quantity=qty, received_qty=0))

    db.commit()
    db.refresh(req)
    return req


@router.post("/{req_id}/items/{req_item_id}/receive")
def receive_item(
    req_id: int,
    req_item_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    req = (
        db.query(Requisition)
        .options(joinedload(Requisition.items))
        .filter(Requisition.id == req_id, Requisition.vessel_id == current_user.vessel_id)
        .first()
    )
    if not req:
        raise HTTPException(404, "Requisition not found")
    if not can_receive(req, current_user):
        raise HTTPException(403, "Receiving not allowed")

    line = next((i for i in req.items if i.id == req_item_id), None)
    if not line:
        raise HTTPException(404, "Item not found")

    qty = data.get("quantity")
    if not qty or qty <= 0:
        raise HTTPException(400, "Invalid quantity")

    remaining = line.quantity - line.received_qty
    if qty > remaining:
        raise HTTPException(400, f"Remaining quantity: {remaining}")

    line.received_qty += qty
    total = sum(i.quantity for i in req.items)
    received = sum(i.received_qty for i in req.items)
    req.status = "received" if received >= total else "partially_received"

    db.commit()
    db.refresh(req)
    return req


@router.delete("/{req_id}")
def delete_requisition(
    req_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    req = get_req_or_404(req_id, current_user.vessel_id, db)
    if not can_delete(req, current_user):
        raise HTTPException(403, "Only captain can delete draft or cancelled requisitions")
    db.delete(req)
    db.commit()
    return {"status": "deleted"}


@router.get("/{req_id}/export")
def export_requisition(
    req_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    req = (
        db.query(Requisition)
        .options(
            joinedload(Requisition.supplier),
            joinedload(Requisition.items).joinedload(RequisitionItem.item),
        )
        .filter(Requisition.id == req_id, Requisition.vessel_id == current_user.vessel_id)
        .first()
    )
    if not req:
        raise HTTPException(404, "Requisition not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Requisition"
    thin = Side(border_style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["B1"] = "Requisition ID"; ws["C1"] = req.id
    ws["B2"] = "Status";          ws["C2"] = req.status
    ws["B3"] = "Vessel";          ws["C3"] = req.vessel.name if req.vessel else ""
    ws["B4"] = "Supplier";        ws["C4"] = req.supplier.name if req.supplier else ""
    ws["B5"] = "Created";         ws["C5"] = req.created_at.strftime("%d-%m-%Y")

    ws.append([])
    ws.append(["Nr", "Item", "Unit", "Qty", "Received", "Description"])
    for cell in ws[7]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for idx, line in enumerate(req.items, start=1):
        ws.append([idx, line.item.name, line.item.unit, line.quantity, line.received_qty, line.item.desc_short or ""])

    for row in ws.iter_rows(min_row=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=requisition_{req.id}.xlsx"},
    )
